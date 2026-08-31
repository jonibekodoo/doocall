"""Calls list filters/pagination/export/columns + dashboard (calls are immutable)."""

from __future__ import annotations

import csv
import io

import pytest
from rest_framework.test import APIClient

from apps.accounts.models import OperatorProfile
from apps.api import storage
from apps.companies.models import Company
from apps.web.models import ExportJob
from apps.web.tasks import run_export

from .fixture_calls import N1, N4, seed_fixture_calls

pytestmark = pytest.mark.django_db

CALLS_URL = "/api/web/v1/calls"


@pytest.fixture
def seeded(
    company: Company,
    op_a: OperatorProfile,
    op_b: OperatorProfile,
    other_company_noise: Company,
) -> dict[str, OperatorProfile]:
    seed_fixture_calls(company, op_a, op_b)
    return {"A": op_a, "B": op_b}


class TestCallsListFilters:
    def test_pagination_totals(self, client: APIClient, seeded: dict) -> None:
        body = client.get(CALLS_URL).json()
        assert body["count"] == 40  # noise tenant excluded
        assert body["pages"] == 2
        assert body["page_size"] == 30
        assert len(body["results"]) == 30
        page2 = client.get(f"{CALLS_URL}?page=2").json()
        assert len(page2["results"]) == 10

    def test_filter_combination(self, client: APIClient, seeded: dict) -> None:
        # operator A + inbound + answered → #1, #11, #19, #33, #35 = 5
        a_id = seeded["A"].pk
        body = client.get(f"{CALLS_URL}?employees={a_id}&direction=inbound&status=answered").json()
        assert body["count"] == 5

    def test_date_range(self, client: APIClient, seeded: dict) -> None:
        body = client.get(f"{CALLS_URL}?date_from=2026-08-04&date_to=2026-08-05").json()
        assert body["count"] == 20  # Tue + Wed

    def test_search_by_number_fragment_and_name(self, client: APIClient, seeded: dict) -> None:
        assert client.get(f"{CALLS_URL}?search=1110004").json()["count"] == 5  # N4
        assert client.get(f"{CALLS_URL}?search=Client 4").json()["count"] == 5

    def test_min_duration(self, client: APIClient, seeded: dict) -> None:
        assert client.get(f"{CALLS_URL}?min_duration=100").json()["count"] == 5
        # ≥100s → #3 (120), #8 (300), #15 (200), #23 (150), #30 (120)

    def test_sort_by_duration(self, client: APIClient, seeded: dict) -> None:
        rows = client.get(f"{CALLS_URL}?ordering=-duration").json()["results"]
        assert rows[0]["duration"] == 300

    def test_sim_filter(self, client: APIClient, seeded: dict) -> None:
        assert client.get(f"{CALLS_URL}?sim_slot=0").json()["count"] == 40
        assert client.get(f"{CALLS_URL}?sim_slot=1").json()["count"] == 0


class TestCallDetailAndDelete:
    def test_detail_includes_audio_urls(
        self, client: APIClient, seeded: dict, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        from apps.calls.models import CallAudio, CallRecord

        record = CallRecord.all_objects.get(call_id="fx-1")
        CallAudio.objects.create(call=record, filename="a.ogg", object_key="k/a.ogg", size_bytes=10)
        monkeypatch.setattr(storage, "presigned_url", lambda key: f"https://s3.test/{key}")
        body = client.get(f"{CALLS_URL}/{record.pk}").json()["call"]
        assert body["audios"] == [
            {
                "kind": "primary",
                "filename": "a.ogg",
                "size_bytes": 10,
                "url": "https://s3.test/k/a.ogg",
            }
        ]

    def test_calls_cannot_be_deleted_by_anyone(
        self, client: APIClient, member_client: APIClient, seeded: dict
    ) -> None:
        """Call records are immutable evidence — the delete endpoint is GONE."""
        from apps.calls.models import CallRecord

        record = CallRecord.all_objects.get(call_id="fx-1")
        assert client.delete(f"{CALLS_URL}/{record.pk}/delete").status_code == 404
        assert member_client.delete(f"{CALLS_URL}/{record.pk}/delete").status_code == 404
        assert client.delete(f"{CALLS_URL}/{record.pk}").status_code == 405
        from apps.calls.models import CallRecord

        assert CallRecord.all_objects.filter(pk=record.pk).exists()


class TestColumnPreferences:
    def test_default_then_custom(self, client: APIClient, company: Company) -> None:
        default = client.get(f"{CALLS_URL}/columns").json()["columns"]
        assert "duration" in default
        custom = ["start_time", "duration"]
        assert (
            client.put(f"{CALLS_URL}/columns", {"columns": custom}, format="json").json()["columns"]
            == custom
        )
        assert client.get(f"{CALLS_URL}/columns").json()["columns"] == custom


class TestExport:
    @pytest.fixture(autouse=True)
    def _capture_storage(self, monkeypatch: pytest.MonkeyPatch) -> dict[str, bytes]:
        stored: dict[str, bytes] = {}

        def fake_store(key: str, payload: bytes, filename: str) -> str:
            stored[key] = payload
            return f"https://s3.test/{key}"

        monkeypatch.setattr(storage, "store_audio", fake_store)
        monkeypatch.setattr(storage, "presigned_url", lambda key: f"https://s3.test/{key}")
        # Run the Celery task inline (no broker in unit tests).
        monkeypatch.setattr(run_export, "delay", lambda job_id: run_export(job_id))
        self.stored = stored
        return stored

    def test_csv_export_produces_valid_file(self, client: APIClient, seeded: dict) -> None:
        response = client.post(
            f"{CALLS_URL}/export",
            {"format": "csv", "filters": {"status": "answered"}},
            format="json",
        )
        assert response.status_code == 202
        export_id = response.json()["export_id"]

        body = client.get(f"{CALLS_URL}/export/{export_id}").json()
        assert body["status"] == "done"
        assert body["row_count"] == 21  # answered calls only
        assert body["url"].startswith("https://s3.test/exports/")

        job = ExportJob.all_objects.get(pk=export_id)
        rows = list(csv.DictReader(io.StringIO(self.stored[job.object_key].decode("utf-8-sig"))))
        assert len(rows) == 21
        assert {r["call_status"] for r in rows} == {"answered"}

    def test_xlsx_export(self, client: APIClient, seeded: dict) -> None:
        from openpyxl import load_workbook

        response = client.post(f"{CALLS_URL}/export", {"format": "xlsx"}, format="json")
        export_id = response.json()["export_id"]
        job = ExportJob.all_objects.get(pk=export_id)
        assert job.status == "done"
        wb = load_workbook(io.BytesIO(self.stored[job.object_key]))
        ws = wb.active
        assert ws.max_row == 41  # header + 40 calls

    def test_bad_format_rejected(self, client: APIClient, company: Company) -> None:
        assert (
            client.post(f"{CALLS_URL}/export", {"format": "pdf"}, format="json").status_code == 400
        )


class TestDashboard:
    def test_shape_and_operator_filter(self, client: APIClient, seeded: dict) -> None:
        # Fixture calls are in the past → use a custom check via reports instead;
        # dashboard "today" window is empty, which is itself a valid assertion.
        body = client.get("/api/web/v1/dashboard?period=today").json()
        assert body["general"]["all"]["total"] == 0

        # 7d window from BASE won't include fixture either (fixed dates), so
        # create one fresh call now and confirm it appears.
        from django.utils import timezone

        from apps.calls.models import CallRecord

        CallRecord.all_objects.create(
            company=seeded["A"].company,
            operator=seeded["A"],
            call_id="fresh-1",
            call_type="inbound",
            call_status="answered",
            from_number=N1,
            to_number="+998990000000",
            counterparty_number=N1,
            duration=10,
            start_time=timezone.now(),
        )
        body = client.get("/api/web/v1/dashboard?period=today").json()
        assert body["general"]["all"]["total"] == 1
        assert body["latest_calls"][0]["call_id"] == "fresh-1"
        a_row = next(r for r in body["per_operator"] if r["user_name"] == "op-a")
        assert a_row["total"] == 1

        filtered = client.get(
            f"/api/web/v1/dashboard?period=today&operator={seeded['B'].pk}"
        ).json()
        assert filtered["general"]["all"]["total"] == 0

    def test_unanswered_now_on_dashboard(self, client: APIClient, seeded: dict) -> None:
        from django.utils import timezone

        from apps.calls.models import CallRecord

        CallRecord.all_objects.create(
            company=seeded["B"].company,
            operator=seeded["B"],
            call_id="fresh-missed",
            call_type="inbound",
            call_status="no_answer",
            from_number=N4,
            to_number="+998990000000",
            counterparty_number=N4,
            duration=0,
            start_time=timezone.now(),
        )
        body = client.get("/api/web/v1/dashboard?period=today").json()
        assert [r["counterparty_number"] for r in body["unanswered_now"]] == [N4]
